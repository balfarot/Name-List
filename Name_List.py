#List of Name out of a Excel Sheet
from openpyxl import workbook, load_workbook
import random
wb = load_workbook('hello.xlsx')
ws = wb.active
rangeline = ws['A2':'A19'] #we are trying to create a list here
#our list is named as a rangline above (we can name it anything)

name = [] #creating a empty variable ready to take LIST

for items in rangeline: #this gave us cell values
    for subitems in items: #inside the cell
        name.append(subitems.value)

computer_action = random.choice(name)
print('The computer randomly chose ' + computer_action)